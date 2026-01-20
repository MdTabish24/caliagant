"""
Audio Tracking Excel Handler
Tracks phone numbers, audio listen time, and color codes based on percentage
"""
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from config import logger


class AudioTracker:
    def __init__(self, excel_path=None):
        if excel_path is None:
            # Save in results folder next to exe/script
            import sys
            if getattr(sys, 'frozen', False):
                # Running as exe
                base_dir = os.path.dirname(sys.executable)
            else:
                # Running as script
                base_dir = os.path.dirname(os.path.abspath(__file__))
            
            results_dir = os.path.join(base_dir, "results")
            os.makedirs(results_dir, exist_ok=True)
            excel_path = os.path.join(results_dir, "audio_tracking.xlsx")
        self.excel_path = os.path.abspath(excel_path)
        logger.info(f"📊 Audio tracker Excel path: {self.excel_path}")
        self._init_excel()
    
    def _init_excel(self):
        """Initialize Excel with headers"""
        if os.path.exists(self.excel_path):
            logger.info(f"📊 Audio tracking Excel exists: {self.excel_path}")
            return
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Audio Tracking"
        
        headers = ["Phone Number", "Date Time", "Audio Length (s)", "Listened (s)", "Percentage", "Status"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        
        wb.save(self.excel_path)
        logger.info(f"📊 Created audio tracking Excel: {self.excel_path}")
    
    def log_call(self, phone_number, audio_length, listened_time):
        """
        Log call with color coding
        RED: < 20% listened
        YELLOW: 20-60% listened
        GREEN: > 60% listened
        """
        try:
            wb = load_workbook(self.excel_path)
            ws = wb.active
            
            # Calculate percentage
            percentage = (listened_time / audio_length * 100) if audio_length > 0 else 0
            
            # Determine status and color
            if percentage < 20:
                status = "NOT INTERESTED"
                fill_color = "FF6B6B"  # RED
            elif percentage < 60:
                status = "PARTIAL"
                fill_color = "FFD93D"  # YELLOW
            else:
                status = "INTERESTED"
                fill_color = "6BCF7F"  # GREEN
            
            # Add row
            row_data = [
                phone_number,
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                round(audio_length, 1),
                round(listened_time, 1),
                f"{percentage:.1f}%",
                status
            ]
            
            ws.append(row_data)
            
            # Apply color to entire row
            last_row = ws.max_row
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            
            for col in range(1, 7):
                cell = ws.cell(row=last_row, column=col)
                cell.fill = fill
                cell.alignment = Alignment(horizontal="center")
            
            wb.save(self.excel_path)
            logger.info(f"📊 Logged: {phone_number} | {percentage:.1f}% | {status}")
            
        except Exception as e:
            logger.error(f"Audio tracking error: {e}")


if __name__ == "__main__":
    tracker = AudioTracker()
    tracker.log_call("9876543210", 30, 5)   # RED
    tracker.log_call("9876543211", 30, 15)  # YELLOW
    tracker.log_call("9876543212", 30, 25)  # GREEN
    print("Test complete!")
