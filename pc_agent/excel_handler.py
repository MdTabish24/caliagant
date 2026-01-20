"""
Excel Handler - Results save karna
"""
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from config import OUTPUT_EXCEL, logger


class ExcelHandler:
    def __init__(self):
        self.output_file = OUTPUT_EXCEL
        logger.info(f"📊 Excel handler output path: {self.output_file}")
        self._init_file()
    
    def _init_file(self):
        """Initialize Excel file with headers"""
        # Always ensure directory exists
        os.makedirs(os.path.dirname(self.output_file) or ".", exist_ok=True)
        
        if os.path.exists(self.output_file):
            logger.info(f"📊 Excel file exists: {self.output_file}")
            return
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Call Results"
        
        headers = ["Phone", "Time", "Duration", "Interest", "Result", "Summary", "Conversation"]
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 40
        ws.column_dimensions['G'].width = 50
        
        wb.save(self.output_file)
        logger.info(f"📊 Created Excel: {self.output_file}")
    
    def save_result(self, phone, duration, analysis, conversation):
        """Save call result"""
        try:
            wb = load_workbook(self.output_file)
            ws = wb.active
            
            row = [
                phone,
                datetime.now().strftime("%Y-%m-%d %H:%M"),
                duration,
                analysis.get("interest", ""),
                analysis.get("result", ""),
                analysis.get("summary", ""),
                conversation[:500] if conversation else ""
            ]
            
            ws.append(row)
            
            # Color code
            last_row = ws.max_row
            result = analysis.get("result", "").upper()
            
            if "POSITIVE" in result:
                fill = PatternFill(start_color="C6EFCE", fill_type="solid")
            elif "NEGATIVE" in result:
                fill = PatternFill(start_color="FFC7CE", fill_type="solid")
            else:
                fill = PatternFill(start_color="FFEB9C", fill_type="solid")
            
            for col in range(1, 8):
                ws.cell(row=last_row, column=col).fill = fill
            
            wb.save(self.output_file)
            logger.info(f"💾 Saved result for {phone}")
            
        except Exception as e:
            logger.error(f"Excel save error: {e}")


if __name__ == "__main__":
    handler = ExcelHandler()
    handler.save_result("9876543210", 45, {
        "interest": "INTERESTED",
        "result": "POSITIVE",
        "summary": "User interested in course"
    }, "User: Hello\nAI: Hello sir")
    print("Done!")
